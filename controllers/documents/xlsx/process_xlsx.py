import threading
from openpyxl import load_workbook
import pandas as pd


def format_percent(value):
    """Định dạng giá trị phần trăm nếu cần."""
    if isinstance(value, float):
        if 0 <= value <= 1:
            return f"{value * 100:.1f}%"
    return value


def safe_strip(value):
    """Chuyển giá trị thành chuỗi và loại bỏ khoảng trắng."""
    if pd.isna(value):
        return ""
    return str(value).strip()


def extract_data(sheet_data):
    sheet_data = sheet_data.iloc[2:]  # Bỏ qua 2 dòng đầu tiên

    processed_data = []

    for index, row in sheet_data.iterrows():
        row_dict = {}
        _row_dict = []
        first_data = ""

        for col in sheet_data.columns:
            value = row[col]

            if isinstance(value, str) and '%' in value:
                row_dict[col] = value
            elif pd.isna(value):
                row_dict[col] = ""
            else:
                row_dict[col] = format_percent(value)

            if str(row_dict[col]).lower() in ["vượt", "đạt", "chưa đạt", "không đạt"]:
                row_dict[col] = ""

            if index == 2 and row_dict[col] == "":
                row_dict[col] = first_data

            # Nếu từ dòng thứ 5 trở đi, thay thế giá trị rỗng bằng số 0
            if index >= 5 and row_dict[col] == "":
                row_dict[col] = 0

            _row_dict.append(row_dict[col])
            first_data = row_dict[col]

        processed_data.append(_row_dict)

    # Chuyển processed_data thành DataFrame để dễ thao tác
    df_processed = pd.DataFrame(processed_data, columns=sheet_data.columns)

    # Loại bỏ các cột toàn bộ đều là giá trị rỗng ngay từ đầu
    df_processed = df_processed.dropna(axis=1, how='all')  # Bỏ cột toàn NaN
    df_processed = df_processed.loc[:, (df_processed != "").any(axis=0)]  # Bỏ cột toàn chuỗi rỗng

    return df_processed.values.tolist()  # Trả về danh sách kết quả


def fill_merged_cells(sheet):
    """Điền giá trị vào các ô bị gộp."""
    merged_ranges = list(sheet.merged_cells.ranges)  # Lưu lại các vùng gộp

    # Tạm thời tách các ô gộp
    for merged_range in merged_ranges:
        sheet.unmerge_cells(str(merged_range))

    # Gán giá trị cho các ô trước đó đã gộp
    for merged_range in merged_ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        top_left_value = sheet.cell(row=min_row, column=min_col).value

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = top_left_value

    # Gộp lại các ô như ban đầu
    for merged_range in merged_ranges:
        sheet.merge_cells(str(merged_range))


def process_xlsx(file_path):
    """Đọc file Excel và chuyển dữ liệu thành JSON."""
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        all_data = {}

        print("Sheet names: ", sheet_names)

        try:
            for sheet_name in sheet_names:
                try:
                    sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', header=None)
                    if sheet_data.empty:
                        continue

                    # Ghép 2 dòng đầu tiên làm tiêu đề
                    title_line1 = safe_strip(sheet_data.iloc[0, 0])
                    title_line2 = safe_strip(sheet_data.iloc[1, 0])
                    title = f"{title_line1} {title_line2}" if title_line2 else title_line1

                    extracted_data = extract_data(sheet_data)
                    all_data[sheet_name] = {
                        "title": title,
                        "datas": extracted_data
                    }

                except Exception as e:
                    print(f"Lỗi khi xử lý sheet {sheet_name}: {e}")
                    continue

        except Exception as e:
            pass

        return all_data

    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")
        return None
